# Файл: app/services/call_export.py

"""Сервис формирования XLSX с расшифровками звонков."""

from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
from typing import Any, Dict, Iterable, List, Tuple
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.db.manager import DatabaseManager
from app.logging_config import get_watchdog_logger

logger = get_watchdog_logger(__name__)

MOSCOW_TZ = ZoneInfo("Europe/Moscow")

EXPORT_PERIOD_OPTIONS = (14, 30, 60, 180)

HEADERS = (
    "Дата",
    "Время",
    "Оператор",
    "Врач",
    "Услуга",
    "Номер, с которого звонили",
    "Номер, на который звонили",
    "Тип звонка",
    "Вход/исход",
    "Длительность разговора (в секундах)",
    "Целевой/нецелевой (0/1)",
    "Расшифровка",
    "Анализ Звонка",
    "Цель звонка",
    "Специальность врача",
    "Исход звонка",
    "Возражение (0/1)",
    "Возражение обработано (0/1)",
    "Попытка записи (0/1)",
    "Следующий шаг ясен (0/1)",
    "Follow-up зафиксирован (0/1)",
    "Причина отказа",
    "Категория причины",
    "Причины отказа",
    "Группа отказа",
    "Источник",
    "Оценка качества (0–10)",
    "Дата/время оценки",
)

CATEGORY_LABELS = {
    "запись на услугу": "запись",
    "лид (без записи)": "лид (без записи)",
    "отмена записи": "отмена",
    "перенос записи": "перенос",
    "напоминание о приеме": "подтверждение",
    "жалоба": "жалоба",
    "навигация": "вход/адрес",
    "информация о состоянии пациента": "результаты/анализы",
    "спам": "спам",
    "спам, реклама": "спам, реклама",
    "технический звонок": "технический",
}

OUTCOME_LABELS = {
    "record": "запись",
    "lead_no_record": "лид (без записи)",
    "info_only": "справка",
    "non_target": "нецелевой",
}

KEYWORD_LABELS: Tuple[Tuple[str, str], ...] = (
    ("времен", "уточнение времени"),
    ("анализ", "результаты/анализы"),
    ("результат", "результаты/анализы"),
    ("перенос", "перенос"),
    ("отмен", "отмена"),
    ("подтвержд", "подтверждение"),
    ("жалоб", "жалоба"),
    ("адрес", "вход/адрес"),
)


class CallExportService:
    """Генератор выгрузок звонков в формате XLSX."""

    def __init__(self, db_manager: DatabaseManager):
        self.db_manager = db_manager

    async def build_export(self, days: int) -> Tuple[BytesIO, str, int, Tuple[datetime, datetime]]:
        """
        Строит XLSX-файл с расшифровками за указанное количество дней.

        Возвращает bytes-поток, имя файла, количество строк и границы периода.
        """

        if days not in EXPORT_PERIOD_OPTIONS:
            raise ValueError(f"Unsupported export period: {days}")

        period_start, period_end = self._resolve_period(days)
        rows = await self._fetch_calls(period_start, period_end)
        logger.info(
            "[CALL_EXPORT] Building workbook: %s rows for %s-%s",
            len(rows),
            period_start,
            period_end,
        )
        workbook = self._build_workbook(rows)
        filename = f"Выгрузка_звонков_{period_start:%Y%m%d}_{period_end:%Y%m%d}.xlsx"
        return workbook, filename, len(rows), (period_start, period_end)

    async def _fetch_calls(
        self,
        start: datetime,
        end: datetime,
    ) -> List[Dict[str, Any]]:
        base_select = [
            "cs.history_id",
            "cs.call_date",
            "cs.called_info",
            "cs.requested_doctor_name",
            "cs.requested_service_name",
            "cs.caller_number",
            "cs.called_number",
            "cs.call_type",
            "cs.context_type",
            "cs.talk_duration",
            "cs.is_target",
            "cs.transcript",
            "cs.result",
            "cs.call_category",
            "cs.requested_doctor_speciality",
            "cs.outcome",
        ]
        optional_columns = [
            "objection_present",
            "objection_handled",
            "booking_attempted",
            "next_step_clear",
            "followup_captured",
        ]
        tail_select = [
            "cs.refusal_reason",
            "rc.label AS refusal_category_label",
            "cs.refusal_category_code",
            "cs.refusal_group",
            "cs.utm_source_by_number",
            "cs.call_score",
            "cs.score_date",
        ]

        columns = await self._get_call_scores_columns()
        if columns is None:
            query = self._build_export_query(
                base_select,
                optional_columns,
                tail_select,
                available_columns=None,
            )
            try:
                result = await self.db_manager.execute_with_retry(
                    query,
                    params=(start, end),
                    fetchall=True,
                    query_name="call_export.fetch_calls",
                )
                return [dict(row) for row in (result or [])]
            except Exception as exc:
                if "Unknown column" not in str(exc):
                    raise
                logger.warning(
                    "[CALL_EXPORT] Missing columns, exporting without them: %s",
                    exc,
                    exc_info=True,
                )
                logger.warning(
                    "[CALL_EXPORT] Отсутствующие колонки: %s",
                    ", ".join(optional_columns),
                )
                query = self._build_export_query(
                    base_select,
                    optional_columns,
                    tail_select,
                    available_columns=set(),
                )
        else:
            missing = [name for name in optional_columns if name not in columns]
            if missing:
                logger.warning(
                    "[CALL_EXPORT] Отсутствующие колонки: %s",
                    ", ".join(missing),
                )
            query = self._build_export_query(
                base_select,
                optional_columns,
                tail_select,
                available_columns=columns,
            )

        result = await self.db_manager.execute_with_retry(
            query,
            params=(start, end),
            fetchall=True,
            query_name="call_export.fetch_calls",
        )
        return [dict(row) for row in (result or [])]

    async def _get_call_scores_columns(self) -> set[str] | None:
        query = """
            SELECT COLUMN_NAME
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME = 'call_scores'
        """
        try:
            rows = await self.db_manager.execute_with_retry(
                query,
                fetchall=True,
                query_name="call_export.call_scores_columns",
            )
        except Exception as exc:
            logger.warning(
                "[CALL_EXPORT] Не удалось получить список колонок call_scores: %s",
                exc,
                exc_info=True,
            )
            return None
        columns: set[str] = set()
        for row in rows or []:
            if isinstance(row, dict):
                name = row.get("COLUMN_NAME")
            else:
                name = row[0] if row else None
            if name:
                columns.add(str(name))
        return columns

    @staticmethod
    def _build_export_query(
        base_select: List[str],
        optional_columns: List[str],
        tail_select: List[str],
        *,
        available_columns: set[str] | None,
    ) -> str:
        select_parts = list(base_select)
        if available_columns is None:
            select_parts.extend([f"cs.{name}" for name in optional_columns])
        else:
            for name in optional_columns:
                if name in available_columns:
                    select_parts.append(f"cs.{name}")
                else:
                    select_parts.append(f"NULL AS {name}")
        select_parts.extend(tail_select)
        select_clause = ",\n                ".join(select_parts)
        return f"""
            SELECT
                {select_clause}
            FROM call_scores cs
            LEFT JOIN refusal_categories rc ON rc.id = cs.refusal_category_id
            WHERE cs.call_date BETWEEN %s AND %s
            ORDER BY cs.call_date ASC
        """

    def _resolve_period(self, days: int) -> Tuple[datetime, datetime]:
        now = datetime.now(MOSCOW_TZ)
        start_date = (now - timedelta(days=max(days - 1, 0))).replace(
            hour=0,
            minute=0,
            second=0,
            microsecond=0,
        )
        end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
        return start_date.replace(tzinfo=None), end_date.replace(tzinfo=None)

    def _build_workbook(self, rows: Iterable[Dict[str, Any]]) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Звонки"

        ws.append(HEADERS)
        header_font = Font(bold=True)
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=1, column=col_idx).font = header_font

        wrap_columns = {12, 13, 14}
        row_index = 2
        for row in rows:
            call_date = row.get("call_date")
            if not isinstance(call_date, datetime):
                continue
            score_date = row.get("score_date")
            goal = self._resolve_goal(row)
            values = [
                call_date.date(),
                call_date.time(),
                self._normalize_label(row.get("called_info")),
                self._normalize_label(row.get("requested_doctor_name")),
                self._normalize_label(row.get("requested_service_name")),
                self._normalize_label(row.get("caller_number")),
                self._normalize_label(row.get("called_number")),
                self._normalize_label(row.get("call_type")),
                self._normalize_label(row.get("context_type")),
                row.get("talk_duration"),
                row.get("is_target"),
                (row.get("transcript") or "").strip(),
                (row.get("result") or "").strip(),
                goal,
                self._normalize_label(row.get("requested_doctor_speciality")),
                self._normalize_label(row.get("outcome")),
                row.get("objection_present"),
                row.get("objection_handled"),
                row.get("booking_attempted"),
                row.get("next_step_clear"),
                row.get("followup_captured"),
                self._normalize_label(row.get("refusal_reason")),
                self._normalize_label(row.get("refusal_category_label")),
                self._normalize_label(row.get("refusal_category_code")),
                self._normalize_label(row.get("refusal_group")),
                self._normalize_label(row.get("utm_source_by_number")) or "не указано",
                row.get("call_score"),
                score_date,
            ]
            for idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_index, column=idx, value=value)
                if idx == 1 and isinstance(value, datetime):
                    cell.number_format = "DD.MM.YYYY"
                elif idx == 1:
                    cell.number_format = "DD.MM.YYYY"
                if idx == 2 and isinstance(value, datetime):
                    cell.number_format = "HH:MM"
                elif idx == 2:
                    cell.number_format = "HH:MM"
                if idx == len(values):
                    cell.number_format = "DD.MM.YYYY HH:MM"
                if idx in wrap_columns:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_index += 1

        column_widths = {
            1: 12,
            2: 10,
            3: 24,
            4: 24,
            5: 26,
            6: 24,
            7: 24,
            8: 18,
            9: 14,
            10: 18,
            11: 20,
            12: 80,
            13: 60,
            14: 26,
            15: 26,
            16: 18,
            17: 20,
            18: 24,
            19: 20,
            20: 20,
            21: 20,
            22: 24,
            23: 18,
            24: 20,
            25: 24,
            26: 20,
            27: 24,
            28: 24,
        }
        for column, width in column_widths.items():
            ws.column_dimensions[get_column_letter(column)].width = width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _resolve_goal(self, row: Dict[str, Any]) -> str:
        category = self._normalize_label(row.get("call_category"))
        if category:
            normalized = self._map_category(category)
            if normalized:
                return normalized

        refusal = self._normalize_label(row.get("refusal_reason"))
        if refusal:
            keyword = self._match_keyword(refusal)
            if keyword:
                return keyword

        service = self._normalize_label(row.get("requested_service_name"))
        if service:
            keyword = self._match_keyword(service)
            if keyword:
                return keyword

        outcome = (row.get("outcome") or "").strip().lower()
        if outcome in OUTCOME_LABELS:
            return OUTCOME_LABELS[outcome]

        if row.get("is_target"):
            return "лид (без записи)"

        return "не указано"

    @staticmethod
    def _normalize_label(value: Any) -> str:
        if not value:
            return ""
        text = str(value).strip()
        return text

    def _map_category(self, category: str) -> str:
        normalized = category.lower()
        for key, label in CATEGORY_LABELS.items():
            if key in normalized:
                return label
        keyword = self._match_keyword(category)
        if keyword:
            return keyword
        return category

    @staticmethod
    def _match_keyword(text: str) -> str:
        normalized = text.lower()
        for fragment, label in KEYWORD_LABELS:
            if fragment in normalized:
                return label
        return ""
