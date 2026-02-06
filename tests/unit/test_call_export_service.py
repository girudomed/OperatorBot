from datetime import datetime

import pytest
from openpyxl import load_workbook

from app.services.call_export import CallExportService, HEADERS


class _DummyDBManager:
    def __init__(self):
        self.last_query = None

    async def execute_with_retry(self, query, **kwargs):
        self.last_query = query
        return []


def test_export_headers_do_not_include_removed_refusal_columns():
    assert "Причины отказа" not in HEADERS
    assert "Группа отказа" not in HEADERS
    assert "Категория причины" in HEADERS


@pytest.mark.asyncio
async def test_fetch_calls_query_does_not_reference_removed_refusal_columns():
    db = _DummyDBManager()
    service = CallExportService(db)

    await service._fetch_calls(datetime(2025, 1, 1), datetime(2025, 1, 2))

    assert db.last_query is not None
    assert "cs.refusal_category_code" not in db.last_query
    assert "cs.refusal_group" not in db.last_query


def test_build_workbook_has_same_number_of_headers_and_data_columns():
    service = CallExportService(_DummyDBManager())
    row = {
        "call_date": datetime(2025, 1, 1, 12, 30, 0),
        "called_info": "Оператор",
        "requested_doctor_name": "Иванов",
        "requested_service_name": "Услуга",
        "caller_number": "79990000000",
        "called_number": "70000000000",
        "call_type": "inbound",
        "context_type": "in",
        "talk_duration": 60,
        "is_target": 1,
        "transcript": "txt",
        "result": "ok",
        "call_category": "Лид (без записи)",
        "requested_doctor_speciality": "Терапевт",
        "outcome": "lead_no_record",
        "objection_present": 0,
        "objection_handled": 0,
        "booking_attempted": 0,
        "next_step_clear": 1,
        "followup_captured": 1,
        "refusal_reason": "нет",
        "refusal_category_label": "прочее",
        "utm_source_by_number": "ads",
        "call_score": 8.5,
        "score_date": datetime(2025, 1, 1, 12, 35, 0),
    }

    buffer = service._build_workbook([row])
    wb = load_workbook(buffer)
    ws = wb.active
    header_values = [ws.cell(row=1, column=i).value for i in range(1, len(HEADERS) + 1)]
    first_data_row = [ws.cell(row=2, column=i).value for i in range(1, len(HEADERS) + 1)]

    assert tuple(header_values) == HEADERS
    assert len(first_data_row) == len(HEADERS)
